"""
Create AR6 R10 Scenario Files for VerveStacks
Generates Scen_Par-AR6_R10.xlsx with IEA baseline data and AR6 scenario drivers
"""

import pandas as pd
from pathlib import Path
import sys

# Add parent directory to path to import VerveStacks modules
sys.path.append(str(Path(__file__).parent.parent))

from excel_manager import ExcelManager


def map_iso_to_r10(iso_code):
    """
    Map ISO country code to AR6 R10 region using VS_mappings.xlsx
    
    Args:
        iso_code (str): Three-letter ISO country code (e.g., 'CHE')
        
    Returns:
        str: R10 region code (e.g., 'R10EUROPE')
    """
    try:
        # Load the mapping from VS_mappings.xlsx
        vs_mappings_file = Path("assumptions/VS_mappings.xlsx")
        
        if not vs_mappings_file.exists():
            raise FileNotFoundError(f"VS_mappings.xlsx not found at {vs_mappings_file}")
        
        # Read the AR6 R10 ISO mapping sheet
        mapping_df = pd.read_excel(vs_mappings_file, sheet_name='ar6r10_iso_mapping')

        # Read fuel prices
        fuel_prices_df = pd.read_excel(vs_mappings_file, sheet_name='fuel_prices')
        fuel_prices_df = fuel_prices_df[fuel_prices_df['iso'] == iso_code.upper()]


        # Find the R10 region for this ISO
        iso_row = mapping_df[mapping_df['iso3'] == iso_code.upper()]
        
        if iso_row.empty:
            raise ValueError(f"ISO code '{iso_code}' not found in AR6 R10 mapping")
        
        r10_region = iso_row['r10_region'].iloc[0]
        print(f"✅ Mapped {iso_code} → {r10_region}")
        
        return r10_region, fuel_prices_df
        
    except Exception as e:
        print(f"❌ Error mapping ISO to R10: {e}")
        raise


def load_iea_electricity_summary(iso_code):
    """
    Load IEA electricity summary data for the specified country
    
    Args:
        iso_code (str): Three-letter ISO country code
        
    Returns:
        pd.DataFrame: IEA electricity data for the country
    """
    try:
        # Look for IEA electricity summary file in scenario_drivers folder
        iea_file = Path("scenario_drivers/iea_electricity_summary_2018_2022.csv")
        
        if not iea_file.exists():
            raise FileNotFoundError(f"IEA electricity summary file not found: {iea_file}")
        
        # Load IEA data
        iea_df = pd.read_csv(iea_file)
        
        # Filter for the specified country
        country_data = iea_df[iea_df['iso'] == iso_code.upper()]
        
        if country_data.empty:
            print(f"⚠️  No IEA data found for {iso_code}")
            # Return empty DataFrame with expected structure
            return pd.DataFrame(columns=['iso', 'year', 'industry_twh', 'buildings_agri_twh', 'nonroad_transport_twh', 'road_transport_twh', 'imports_twh', 'exports_twh', 'total_production_twh'])
        
        print(f"✅ Loaded IEA data for {iso_code}: {len(country_data)} records")
        return country_data
        
    except Exception as e:
        print(f"❌ Error loading IEA data: {e}")
        # Return empty DataFrame to allow processing to continue
        return pd.DataFrame(columns=['iso', 'year', 'industry_twh', 'buildings_agri_twh', 'nonroad_transport_twh', 'road_transport_twh', 'imports_twh', 'exports_twh', 'total_production_twh'])


def load_ar6_scenario_drivers(r10_region):
    """
    Load AR6 scenario drivers for the specified R10 region
    
    Args:
        r10_region (str): R10 region code (e.g., 'R10EUROPE')
        
    Returns:
        pd.DataFrame: AR6 scenario data for the region
    """
    try:
        # Load the AR6 scenario drivers we created earlier
        ar6_file = Path("scenario_drivers/ar6_r10_scenario_drivers.csv")
        
        if not ar6_file.exists():
            raise FileNotFoundError(f"AR6 scenario drivers file not found: {ar6_file}")
        
        # Load AR6 data
        ar6_df = pd.read_csv(ar6_file)
        
        # Filter for the specified R10 region
        region_data = ar6_df[ar6_df['Region'] == r10_region]
        
        if region_data.empty:
            raise ValueError(f"No AR6 data found for region '{r10_region}'")
        
        print(f"✅ Loaded AR6 data for {r10_region}: {len(region_data)} records")
        return region_data
        
    except Exception as e:
        print(f"❌ Error loading AR6 data: {e}")
        raise


def create_ar6_r10_scenario(iso_code, grid_modeling=False, data_source=None, output_dir=None):
    """
    Create AR6 R10 scenario file for the specified country
    
    Args:
        iso_code (str): Three-letter ISO country code (e.g., 'CHE')
        output_dir (Path, optional): Output directory. If None, saves to current directory.
        
    Returns:
        dict: Statistics about the generated scenario file
    """
    
    print(f"\n🚀 Creating AR6 R10 scenario for {iso_code}")
    print("=" * 50)
    
    try:
        # Step 1: Map ISO to R10 region
        r10_region, fuel_prices_df = map_iso_to_r10(iso_code)
        
        # Step 2: Load IEA baseline data
        iea_data = load_iea_electricity_summary(iso_code)
        
        # Step 3: Load AR6 scenario drivers
        ar6_data = load_ar6_scenario_drivers(r10_region)
        
        # Step 4: Determine output file path
        if output_dir:
            # Production mode - save to proper VEDA model folder
            suppxls_folder = Path(output_dir) / 'SuppXLS'
            output_file = suppxls_folder / 'Scen_Par-AR6_R10.xlsx'
        else:
            # Testing mode - use template in VerveStacks root
            output_file = Path('Scen_Par-AR6_R10.xlsx')
        
        # Step 5: Update existing Excel template with professional formatting
        print(f"📊 Updating Excel template: {output_file}")
        
        if not output_file.exists():
            raise FileNotFoundError(f"Excel template not found: {output_file}")
        
        # Use ExcelManager with xlwings for professional formatting
        import xlwings as xw
        from excel_manager import ExcelManager
        
        # Initialize trade records counter
        trade_records = 0
        
        # Try professional formatting first, fall back to basic if needed
        formatting_success = False
        
        try:
            # Initialize Excel manager
            excel_manager = ExcelManager()
            
            # Start Excel application and open the workbook
            app = xw.App(visible=False)
            wb = app.books.open(str(output_file))
            
            # Update iea_data sheet with professional formatting
            if 'iea_data' in [sheet.name for sheet in wb.sheets]:
                ws_iea = wb.sheets['iea_data']
                ws_iea.clear()
                excel_manager.write_formatted_table(ws_iea, "A10", iea_data, "IEA Baseline Data")
                print(f"   ✅ Updated 'iea_data' sheet: {len(iea_data)} records (professional formatting)")
            else:
                print(f"   ⚠️  'iea_data' sheet not found in template")
            
            # Update ar6_r10 sheet with professional formatting
            if 'ar6_r10' in [sheet.name for sheet in wb.sheets]:
                ws_ar6 = wb.sheets['ar6_r10']
                ws_ar6.clear()
                excel_manager.write_formatted_table(ws_ar6, "A10", ar6_data, "AR6 Scenario Drivers")
                print(f"   ✅ Updated 'ar6_r10' sheet: {len(ar6_data)} records (professional formatting)")
            else:
                print(f"   ⚠️  'ar6_r10' sheet not found in template")
            
            # Update fuel_prices sheet with professional formatting
            if 'fuel_prices' in [sheet.name for sheet in wb.sheets]:
                ws_fuel_prices = wb.sheets['fuel_prices']
                ws_fuel_prices.clear()
                excel_manager.write_formatted_table(ws_fuel_prices, "A10", fuel_prices_df, "Fuel Prices")
                print(f"   ✅ Updated 'fuel_prices' sheet: {len(fuel_prices_df)} records (professional formatting)")
            else:
                print(f"   ⚠️  'fuel_prices' sheet not found in template")
            
            ws_veda = wb.sheets['Veda']

            if grid_modeling and not data_source.startswith('syn'):
                ws_veda.range('G1').api.NumberFormat = "@"
                ws_veda.range('G1').value = '1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,21,22,23,24,25'
                print(f"   ✅ activated 20 scenarios in parscen")
            else:
                ws_veda.range('G1').api.NumberFormat = "@"
                ws_veda.range('G1').value = '1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,21,22,23,24,25,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,56,57,58,59,60,66,67,68,69,70'
                print(f"   ✅ activated 50 scenarios in parscen")

            # Add electricity trade capacity data to Veda sheet
            print("⚡ Adding electricity trade capacity data...")
            try:
                trade_df = excel_manager.create_electricity_trade_capacity(iso_code)
                
                if not trade_df.empty:

                    # Find the last used row in the Veda sheet
                    try:
                        used_range = ws_veda.used_range
                        if used_range is not None:
                            last_used_row = used_range.last_cell.row
                        else:
                            last_used_row = 0
                    except:
                        # Fallback: manually check for last used row
                        last_used_row = 0
                        for row in range(1, 101):  # Check first 100 rows
                            try:
                                if (ws_veda.range(f'A{row}').value is not None or 
                                    ws_veda.range(f'B{row}').value is not None or
                                    ws_veda.range(f'C{row}').value is not None):
                                    last_used_row = row
                            except:
                                break
                    
                    # Write trade data one row after the last used row
                    current_row = last_used_row + 5
                    excel_manager.write_formatted_table(
                        ws_veda, 
                        f'A{current_row}', 
                        trade_df, 
                        veda_marker="~tfm_ins-ts"
                    )
                    print(f"   ✅ Electricity trade data written to Veda sheet: {len(trade_df)} records")
                    trade_records = len(trade_df)
                else:
                    print("   ℹ️  No electricity interconnections found for this country")
                    trade_records = 0
                    
            except Exception as e:
                print(f"   ⚠️  Error adding electricity trade data: {e}")
                trade_records = 0
            

            # Save and close
            wb.save()
            wb.close()
            app.quit()
            formatting_success = True
            
        except Exception as e:
            print(f"   ❌ Professional formatting failed: {e}")
            print(f"   🔄 Falling back to basic Excel writing...")
            
            # Close any open xlwings workbook and app
            try:
                if 'wb' in locals():
                    wb.close()
                if 'app' in locals():
                    app.quit()
            except:
                pass
        
        # Fallback to basic openpyxl if professional formatting failed
        if not formatting_success:
            from openpyxl import load_workbook
            
            wb = load_workbook(output_file)
            
            # Update iea_data sheet (basic)
            if 'iea_data' in wb.sheetnames:
                ws_iea = wb['iea_data']
                ws_iea.delete_rows(1, ws_iea.max_row)
                for c_idx, col_name in enumerate(iea_data.columns, 1):
                    ws_iea.cell(row=1, column=c_idx, value=col_name)
                for r_idx, row in enumerate(iea_data.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_iea.cell(row=r_idx, column=c_idx, value=value)
                print(f"   ✅ Updated 'iea_data' sheet: {len(iea_data)} records (basic formatting)")
            
            # Update ar6_r10 sheet (basic)
            if 'ar6_r10' in wb.sheetnames:
                ws_ar6 = wb['ar6_r10']
                ws_ar6.delete_rows(1, ws_ar6.max_row)
                for c_idx, col_name in enumerate(ar6_data.columns, 1):
                    ws_ar6.cell(row=1, column=c_idx, value=col_name)
                for r_idx, row in enumerate(ar6_data.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_ar6.cell(row=r_idx, column=c_idx, value=value)
                print(f"   ✅ Updated 'ar6_r10' sheet: {len(ar6_data)} records (basic formatting)")
            
            wb.save(output_file)
            wb.close()
        
        # Step 6: Generate visualization charts for README
        try:
            print("📊 Generating AR6 scenario visualization charts...")
            create_ar6_scenario_charts(r10_region, ar6_data, iso_code)
            print("   ✅ Charts saved to scenario_drivers folder")
            
            # Copy AR6 scenario materials to model folder
            if output_dir:
                copy_ar6_scenario_materials(iso_code, suppxls_folder)
                print("   ✅ AR6 scenario materials copied to model folder")
        except Exception as e:
            print(f"   ⚠️  Could not generate charts: {e}")
        
        # Step 7: Collect statistics for README
        scenario_stats = {
            'iso_code': iso_code,
            'r10_region': r10_region,
            'output_file': str(output_file),
            'iea_records': len(iea_data),
            'ar6_records': len(ar6_data),
            'trade_records': trade_records,
            'scenario_categories': sorted(ar6_data['Category'].unique()) if 'Category' in ar6_data.columns else [],
            'scenario_attributes': sorted(ar6_data['attribute'].unique()) if 'attribute' in ar6_data.columns else [],
            'years_covered': sorted(ar6_data['Year'].unique()) if 'Year' in ar6_data.columns else []
        }
        
        print(f"\n✅ Successfully created AR6 scenario file!")
        print(f"📁 File: {output_file}")
        print(f"🌍 Region: {r10_region}")
        print(f"📊 IEA records: {len(iea_data)}")
        print(f"📈 AR6 records: {len(ar6_data)}")
        print(f"⚡ Trade records: {trade_records}")
        print(f"🎭 Categories: {scenario_stats['scenario_categories']}")
        print(f"📅 Years: {scenario_stats['years_covered']}")
        
        return scenario_stats
        
    except Exception as e:
        print(f"❌ Error creating AR6 scenario: {e}")
        raise


def create_ar6_scenario_charts(r10_region, ar6_data, iso_code):
    """
    Generate three visualization charts for AR6 scenarios:
    1. CO2 Price Trajectories
    2. Electricity Demand Growth (absolute TWh)
    3. Hydrogen Demand (absolute TWh)
    
    Args:
        r10_region: R10 region code (e.g., 'R10EUROPE')
        ar6_data: Filtered AR6 data for the region
        iso_code: Country ISO code for chart titles
    """
    import matplotlib.pyplot as plt
    import numpy as np
    from pathlib import Path
    
    # Create output directory in scenario_drivers
    output_dir = Path("scenario_drivers")
    output_dir.mkdir(exist_ok=True)
    
    # Define climate categories and colors
    categories = ['C1', 'C2', 'C3', 'C4', 'C7']
    colors = {
        'C1': '#2E8B57',  # Sea Green (ambitious)
        'C2': '#4682B4',  # Steel Blue
        'C3': '#DAA520',  # Goldenrod
        'C4': '#CD853F',  # Peru
        'C7': '#B22222'   # Fire Brick (limited action)
    }
    
    # Years for plotting
    years = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
    
    # Load IEA baseline data for absolute calculations and historical plotting
    try:
        iea_file = Path("scenario_drivers/iea_electricity_summary_2018_2022.csv")
        if iea_file.exists():
            import pandas as pd
            iea_df = pd.read_csv(iea_file)
            iso_iea = iea_df[iea_df['iso'] == iso_code.upper()]
            if not iso_iea.empty:
                # Use total electricity production as baseline (most recent year)
                iea_baseline_twh = iso_iea['total_production_twh'].iloc[-1]
                
                # Prepare historical data for plotting
                iea_historical = iso_iea.sort_values('year')
                iea_years = iea_historical['year'].tolist()
                iea_values = iea_historical['total_production_twh'].tolist()
                print(f"   📊 IEA historical data: {len(iea_years)} years ({min(iea_years)}-{max(iea_years)})")
            else:
                iea_baseline_twh = 100  # Fallback for missing data
                iea_years = []
                iea_values = []
        else:
            iea_baseline_twh = 100  # Fallback
            iea_years = []
            iea_values = []
    except Exception as e:
        print(f"   ⚠️  Could not load IEA baseline: {e}")
        iea_baseline_twh = 100
        iea_years = []
        iea_values = []
    
    # Set up the plotting style for delicate, refined visualization
    plt.style.use('default')
    fig_width, fig_height = 18, 6  # Landscape format for three panels
    
    # Create figure with three side-by-side subplots
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(fig_width, fig_height))
    
    # Get country name using data_utils function
    from data_utils import get_country_name_from_iso
    country_name = get_country_name_from_iso(iso_code)
    
    # Calculate scenario count range across categories for this region
    if 'count' in ar6_data.columns:
        min_count = ar6_data['count'].min()
        max_count = ar6_data['count'].max()
        if min_count == max_count:
            scenario_info = f'[median values from {min_count} scenarios each]'
        else:
            scenario_info = f'[median values from {min_count}-{max_count} scenarios]'
    else:
        scenario_info = '[median values]'
    
    fig.suptitle(f'{country_name} ({r10_region}) {scenario_info}', 
                 fontsize=16, fontweight='bold', y=0.95)
    
    # Prepare data for all three charts
    co2_data = ar6_data[ar6_data['attribute'] == 'CO2 price']
    elec_data = ar6_data[ar6_data['attribute'] == 'Electricity growth relative to 2020']
    hydrogen_data = ar6_data[ar6_data['attribute'] == 'Hydrogen as a share of electricity']
    
    # Chart 1: CO2 Price Trajectories (delicate thin lines)
    for category in categories:
        cat_data = co2_data[co2_data['Category'] == category]
        if not cat_data.empty:
            # Get median values for each year
            medians = []
            for year in years:
                year_data = cat_data[cat_data['Year'] == year]
                if not year_data.empty:
                    medians.append(year_data['median'].iloc[0])
                else:
                    medians.append(np.nan)
            
            ax1.plot(years, medians, color=colors[category], linewidth=1.5, 
                    label=f'{category}', alpha=0.8)
    
    ax1.set_title('CO2 Price Trajectories', fontsize=12, fontweight='bold', pad=15)
    ax1.set_ylabel('CO2 Price ($/tCO2)', fontsize=10)
    ax1.grid(True, alpha=0.2, linewidth=0.5)
    ax1.legend(fontsize=9, frameon=False)
    
    # Chart 2: Electricity Demand Growth (delicate thin lines)
    for category in categories:
        cat_data = elec_data[elec_data['Category'] == category]
        if not cat_data.empty:
            # Calculate absolute electricity demand
            absolute_demand = []
            for year in years:
                year_data = cat_data[cat_data['Year'] == year]
                if not year_data.empty:
                    growth_index = year_data['median'].iloc[0]
                    absolute_twh = iea_baseline_twh * growth_index
                    absolute_demand.append(absolute_twh)
                else:
                    absolute_demand.append(np.nan)
            
            ax2.plot(years, absolute_demand, color=colors[category], linewidth=1.5,
                    label=f'{category}', alpha=0.8)
    
    # Add IEA historical data line to electricity demand chart
    if iea_years and iea_values:
        ax2.plot(iea_years, iea_values, color='black', linewidth=2.0, 
                linestyle='-', label='IEA Historical', alpha=0.9, zorder=10)
    
    # Add vertical line at 2020 to separate historical from projections
    ax2.axvline(x=2020, color='gray', linestyle='--', linewidth=1.0, alpha=0.7, zorder=5)
    
    ax2.set_title('Electricity Demand Trajectories', fontsize=12, fontweight='bold', pad=15)
    ax2.set_ylabel('Electricity Demand (TWh)', fontsize=10)
    ax2.set_ylim(bottom=0)  # Start y-axis at 0
    ax2.grid(True, alpha=0.2, linewidth=0.5)
    ax2.legend(fontsize=9, frameon=False)
    
    # Chart 3: Hydrogen Demand (delicate thin lines)
    for category in categories:
        cat_elec = elec_data[elec_data['Category'] == category]
        cat_h2 = hydrogen_data[hydrogen_data['Category'] == category]
        
        if not cat_elec.empty and not cat_h2.empty:
            # Calculate absolute hydrogen demand
            hydrogen_demand = []
            for year in years:
                elec_year = cat_elec[cat_elec['Year'] == year]
                h2_year = cat_h2[cat_h2['Year'] == year]
                
                if not elec_year.empty and not h2_year.empty:
                    growth_index = elec_year['median'].iloc[0]
                    h2_share = h2_year['median'].iloc[0]
                    absolute_elec_twh = iea_baseline_twh * growth_index
                    hydrogen_twh = absolute_elec_twh * h2_share
                    hydrogen_demand.append(hydrogen_twh)
                else:
                    hydrogen_demand.append(np.nan)
            
            ax3.plot(years, hydrogen_demand, color=colors[category], linewidth=1.5,
                    label=f'{category}', alpha=0.8)
    
    ax3.set_title('Hydrogen Demand Trajectories', fontsize=12, fontweight='bold', pad=15)
    ax3.set_ylabel('Hydrogen Demand (TWh)', fontsize=10)
    ax3.grid(True, alpha=0.2, linewidth=0.5)
    ax3.legend(fontsize=9, frameon=False)
    
    # Adjust layout and styling
    plt.tight_layout()
    plt.subplots_adjust(top=0.85, wspace=0.3)  # Space for title
    
    # Save the combined chart
    chart_path = output_dir / f'ar6_scenarios_{iso_code}.png'
    plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"   📊 Combined AR6 scenarios chart: {chart_path}")


def copy_ar6_scenario_materials(iso_code, model_output_dir):
    """
    Copy AR6 scenario visualization materials to the model folder for README embedding
    
    Args:
        iso_code (str): Country ISO code
        model_output_dir (Path): Model output directory (e.g., SuppXLS folder)
    """
    import shutil
    
    # Create scenario_drivers subfolder in model directory
    model_scenario_dir = model_output_dir.parent / "scenario_drivers"
    model_scenario_dir.mkdir(exist_ok=True)
    
    # Source files in VerveStacks scenario_drivers folder
    source_dir = Path("scenario_drivers")
    
    # Copy AR6 scenario chart
    chart_filename = f"ar6_scenarios_{iso_code}.png"
    source_chart = source_dir / chart_filename
    if source_chart.exists():
        dest_chart = model_scenario_dir / chart_filename
        shutil.copy2(source_chart, dest_chart)
        print(f"   ✅ Copied AR6 chart: {chart_filename}")
    else:
        print(f"   ⚠️  AR6 chart not found: {source_chart}")
    
    # Copy any other AR6-related files if they exist
    ar6_files = list(source_dir.glob(f"*{iso_code}*ar6*"))
    for ar6_file in ar6_files:
        if ar6_file.is_file():
            dest_file = model_scenario_dir / ar6_file.name
            shutil.copy2(ar6_file, dest_file)
            print(f"   ✅ Copied AR6 file: {ar6_file.name}")
    
    print(f"   ✅ AR6 scenario materials copied to: {model_scenario_dir}")


def main():
    """Test the AR6 scenario creation with sample countries"""
    
    # Test countries from different R10 regions
    test_countries = ['JPN','USA','DEU','FRA','GBR','ITA','ESP','CAN','AUS']  # Europe, Europe, North America
    
    for iso_code in test_countries:
        try:
            stats = create_ar6_r10_scenario(iso_code)
            print(f"\n🎯 {iso_code} completed successfully!")
        except Exception as e:
            print(f"\n❌ {iso_code} failed: {e}")
        print("-" * 50)


if __name__ == "__main__":
    main()
